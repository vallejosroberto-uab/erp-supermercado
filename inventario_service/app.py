import os
import re
import sqlite3
import uuid
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import requests  # Para la simulación de eventos HTTP al servicio de notificaciones


app = Flask(__name__)
CORS(app)

# =========================================================================
# REGLA #1: RUTA CORREGIDA PARA APUNTAR A LA CARPETA GLOBAL CENTRALIZADA
# =========================================================================
# os.path.dirname(__file__) es 'inventario_service'. 
# El primer dirname() nos sube a la raíz del proyecto 'erp-supermercado'.
DATABASE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), 
    'database', 
    'inventario.db'
)

NOTIFICATION_SERVICE_URL = "http://127.0.0.1:5005/api/notifications"

def get_db_connection():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

# ==========================================
# REGLA #2: VALIDACIONES Y SEGURIDAD
# ==========================================
def es_invalido(campo):
    """Retorna True si el campo es nulo, vacío o contiene solo espacios en blanco."""
    if campo is None or str(campo).strip() == "":
        return True
    return False

def sanitizar_xss(texto):
    if isinstance(texto, str):
        return re.sub(r'<[^>]*>', '', texto)
    return texto

def convertir_entero_positivo(valor, nombre_campo):
    """
    Convierte un valor a entero positivo.
    Rechaza valores vacíos, letras, decimales, negativos, cero y booleanos.
    """
    if isinstance(valor, bool) or es_invalido(valor):
        return None, f"El campo '{nombre_campo}' es obligatorio y debe ser un número entero positivo."

    try:
        if isinstance(valor, float):
            if not valor.is_integer():
                return None, f"El campo '{nombre_campo}' debe ser un número entero, no decimal."
            numero = int(valor)
        else:
            numero = int(str(valor).strip())
    except (ValueError, TypeError):
        return None, f"El campo '{nombre_campo}' debe ser un número entero positivo."

    if numero <= 0:
        return None, f"El campo '{nombre_campo}' debe ser mayor a cero."

    return numero, None

def disparar_evento_notificacion(evento_tipo, cliente, contenido):
    """Simula el envío de eventos HTTP dirigidos al servicio de notificaciones (Regla #3)."""
    payload = {
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "cliente": cliente,
        "tipo": evento_tipo,
        "contenido": contenido
    }
    try:
        requests.post(NOTIFICATION_SERVICE_URL, json=payload, timeout=1)
    except Exception:
        pass

# ==========================================
# ENDPOINTS OBLIGATORIOS DEL MÓDULO
# ==========================================

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        "status": "success",
        "message": "Inventario OK en puerto 5002",
        "ruta_db_usada": DATABASE_PATH
    }), 200


@app.route('/debug-db', methods=['GET'])
def debug_database():
    """Endpoint para verificar que las tablas siguen leyéndose correctamente en la nueva ruta."""
    try:
        conn = get_db_connection()
        tablas_filas = conn.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
        tablas = [f['name'] for f in tablas_filas if f['name'] != 'sqlite_sequence']
        
        estructura_completa = {}
        for tabla in tablas:
            columnas_filas = conn.execute(f"PRAGMA table_info({tabla});").fetchall()
            columnas = [{"columna": c['name'], "tipo": c['type']} for c in columnas_filas]
            estructura_completa[tabla] = {"columnas": columnas}
            
        conn.close()
        return jsonify({
            "status": "success",
            "mensaje": "Base de datos leída con éxito en la nueva ruta global",
            "tablas_encontradas": estructura_completa
        }), 200
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error al leer en la nueva ruta: {str(e)}"}), 500


@app.route('/inventory/balance', methods=['GET'])
def get_balance():
    try:
        conn = get_db_connection()

        query = """
            SELECT 
                i.id,
                i.producto_id,
                p.nombre AS producto_nombre,
                p.codigo AS producto_codigo,
                i.producto_uid,
                i.sucursal_id,
                i.sucursal_uid,
                i.stock_actual,
                i.stock_reservado,
                i.stock_minimo
            FROM inventario i
            JOIN productos p ON i.producto_id = p.id
            ORDER BY p.nombre ASC, i.sucursal_id ASC
        """

        filas = conn.execute(query).fetchall()
        conn.close()

        return jsonify({
            "status": "success",
            "data": [dict(f) for f in filas]
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": "Error al consultar el balance de inventario.",
            "detail": str(e)
        }), 500


@app.route('/inventory/<product>', methods=['GET'])
def get_inventory_by_product(product):
    if es_invalido(product):
        return jsonify({"status": "error", "message": "Identificador de producto inválido."}), 400
    try:
        conn = get_db_connection()
        query = """
            SELECT i.*, p.nombre, p.codigo 
            FROM inventario i
            JOIN productos p ON i.producto_id = p.id
            WHERE p.id = ? OR p.uid = ? OR p.codigo = ?
        """
        filas = conn.execute(query, (product, product, product)).fetchall()
        conn.close()
        return jsonify({"status": "success", "data": [dict(f) for f in filas]}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/inventory/input', methods=['POST'])
def inventory_input():
    data = request.get_json() or {}

    for campo in ['producto_id', 'producto_uid', 'sucursal_id', 'sucursal_uid', 'cantidad']:
        if campo not in data or es_invalido(data[campo]):
            return jsonify({
                "status": "error",
                "message": f"El campo '{campo}' es obligatorio."
            }), 400

    p_id, error = convertir_entero_positivo(data['producto_id'], 'producto_id')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    s_id, error = convertir_entero_positivo(data['sucursal_id'], 'sucursal_id')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    cantidad, error = convertir_entero_positivo(data['cantidad'], 'cantidad')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    p_uid = sanitizar_xss(str(data['producto_uid']).strip())
    s_uid = sanitizar_xss(str(data['sucursal_uid']).strip())

    if es_invalido(p_uid) or es_invalido(s_uid):
        return jsonify({
            "status": "error",
            "message": "Los campos producto_uid y sucursal_uid no pueden contener etiquetas HTML o valores vacíos."
        }), 400

    conn = None

    try:
        conn = get_db_connection()

        registro = conn.execute(
            """
            SELECT stock_actual
            FROM inventario
            WHERE producto_id = ?
            AND sucursal_id = ?
            """,
            (p_id, s_id)
        ).fetchone()

        stock_anterior = 0

        if registro:
            stock_anterior = registro['stock_actual']
            stock_nuevo = stock_anterior + cantidad

            conn.execute(
                """
                UPDATE inventario
                SET stock_actual = ?,
                    actualizado_en = datetime('now')
                WHERE producto_id = ?
                AND sucursal_id = ?
                """,
                (stock_nuevo, p_id, s_id)
            )
        else:
            stock_nuevo = cantidad

            conn.execute(
                """
                INSERT INTO inventario
                (
                    producto_id,
                    producto_uid,
                    sucursal_id,
                    sucursal_uid,
                    stock_actual,
                    stock_reservado,
                    stock_minimo,
                    actualizado_en
                )
                VALUES (?, ?, ?, ?, ?, 0, 5, datetime('now'))
                """,
                (p_id, p_uid, s_id, s_uid, stock_nuevo)
            )

        conn.execute(
            """
            INSERT INTO movimientos_inventario
            (
                uid,
                producto_id,
                producto_uid,
                sucursal_id,
                sucursal_uid,
                tipo_movimiento,
                cantidad,
                stock_anterior,
                stock_nuevo,
                creado_en
            )
            VALUES (?, ?, ?, ?, ?, 'entrada_compra', ?, ?, ?, datetime('now'))
            """,
            (
                str(uuid.uuid4()),
                p_id,
                p_uid,
                s_id,
                s_uid,
                cantidad,
                stock_anterior,
                stock_nuevo
            )
        )

        conn.commit()

        disparar_evento_notificacion(
            "InventoryUpdated",
            "Sistema Inventario",
            f"Ingreso de {cantidad} unidades registrado en sucursal {s_uid}."
        )

        return jsonify({
            "status": "success",
            "message": "Ingreso registrado correctamente.",
            "stock_anterior": stock_anterior,
            "stock_actual": stock_nuevo
        }), 200

    except Exception as e:
        if conn:
            conn.rollback()

        return jsonify({
            "status": "error",
            "message": "Error al registrar ingreso de inventario.",
            "detail": str(e)
        }), 500

    finally:
        if conn:
            conn.close()

@app.route('/inventory/output', methods=['POST'])
def inventory_output():
    data = request.get_json() or {}

    for campo in ['producto_id', 'sucursal_id', 'cantidad']:
        if campo not in data or es_invalido(data[campo]):
            return jsonify({
                "status": "error",
                "message": f"El campo '{campo}' es requerido."
            }), 400

    p_id, error = convertir_entero_positivo(data['producto_id'], 'producto_id')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    s_id, error = convertir_entero_positivo(data['sucursal_id'], 'sucursal_id')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    cantidad, error = convertir_entero_positivo(data['cantidad'], 'cantidad')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    motivo = sanitizar_xss(str(data.get('motivo', 'baja_perdida')).strip())

    if es_invalido(motivo):
        motivo = "baja_perdida"

    conn = None

    try:
        conn = get_db_connection()

        registro = conn.execute(
            """
            SELECT *
            FROM inventario
            WHERE producto_id = ?
            AND sucursal_id = ?
            """,
            (p_id, s_id)
        ).fetchone()

        if not registro:
            return jsonify({
                "status": "error",
                "message": "No existe inventario para el producto y sucursal seleccionados."
            }), 404

        if registro['stock_actual'] < cantidad:
            return jsonify({
                "status": "error",
                "message": "Stock insuficiente para la baja.",
                "stock_disponible": registro['stock_actual']
            }), 400

        stock_anterior = registro['stock_actual']
        stock_nuevo = stock_anterior - cantidad

        conn.execute(
            """
            UPDATE inventario
            SET stock_actual = ?,
                actualizado_en = datetime('now')
            WHERE id = ?
            """,
            (stock_nuevo, registro['id'])
        )

        conn.execute(
            """
            INSERT INTO movimientos_inventario
            (
                uid,
                producto_id,
                producto_uid,
                sucursal_id,
                sucursal_uid,
                tipo_movimiento,
                cantidad,
                stock_anterior,
                stock_nuevo,
                referencia_tipo,
                creado_en
            )
            VALUES (?, ?, ?, ?, ?, 'baja_perdida', ?, ?, ?, ?, datetime('now'))
            """,
            (
                str(uuid.uuid4()),
                p_id,
                registro['producto_uid'],
                s_id,
                registro['sucursal_uid'],
                cantidad,
                stock_anterior,
                stock_nuevo,
                motivo
            )
        )

        conn.commit()

        disparar_evento_notificacion(
            "InventoryUpdated",
            "Sistema Inventario",
            f"Baja de {cantidad} unidades registrada por motivo: {motivo}."
        )

        return jsonify({
            "status": "success",
            "message": "Baja procesada con éxito.",
            "stock_anterior": stock_anterior,
            "stock_actual": stock_nuevo
        }), 200

    except Exception as e:
        if conn:
            conn.rollback()

        return jsonify({
            "status": "error",
            "message": "Error al procesar baja de inventario.",
            "detail": str(e)
        }), 500

    finally:
        if conn:
            conn.close()

    data = request.get_json() or {}
    for campo in ['producto_id', 'sucursal_id', 'cantidad']:
        if campo not in data or es_invalido(data[campo]):
            return jsonify({"status": "error", "message": f"El campo '{campo}' es requerido."}), 400

    cantidad = int(data['cantidad'])
    p_id = data['producto_id']
    s_id = data['sucursal_id']

    try:
        conn = get_db_connection()
        registro = conn.execute("SELECT * FROM inventario WHERE producto_id = ? AND sucursal_id = ?", (p_id, s_id)).fetchone()
        if not registro or registro['stock_actual'] < cantidad:
            conn.close()
            return jsonify({"status": "error", "message": "Stock insuficiente para la baja."}), 400
            
        stock_anterior = registro['stock_actual']
        stock_nuevo = stock_anterior - cantidad
        conn.execute("UPDATE inventario SET stock_actual = ?, actualizado_en = datetime('now') WHERE id = ?", (stock_nuevo, registro['id']))
        conn.execute("INSERT INTO movimientos_inventario (uid, producto_id, producto_uid, sucursal_id, sucursal_uid, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, creado_en) VALUES (?, ?, ?, ?, ?, 'baja_perdida', ?, ?, ?, datetime('now'))", (str(uuid.uuid4()), p_id, registro['producto_uid'], s_id, registro['sucursal_uid'], cantidad, stock_anterior, stock_nuevo))
        conn.commit()
        conn.close()
        return jsonify({"status": "success", "message": "Baja procesada con éxito."}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/inventory/transfer', methods=['POST'])
def inventory_transfer():
    data = request.get_json() or {}

    for campo in ['producto_id', 'sucursal_origen_id', 'sucursal_destino_id', 'cantidad']:
        if campo not in data or es_invalido(data[campo]):
            return jsonify({
                "status": "error",
                "message": f"Falta el campo obligatorio: {campo}"
            }), 400

    p_id, error = convertir_entero_positivo(data['producto_id'], 'producto_id')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    s_origen, error = convertir_entero_positivo(data['sucursal_origen_id'], 'sucursal_origen_id')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    s_destino, error = convertir_entero_positivo(data['sucursal_destino_id'], 'sucursal_destino_id')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    cantidad, error = convertir_entero_positivo(data['cantidad'], 'cantidad')
    if error:
        return jsonify({"status": "error", "message": error}), 400

    if s_origen == s_destino:
        return jsonify({
            "status": "error",
            "message": "La sucursal de origen y destino deben ser distintas."
        }), 400

    conn = None

    try:
        conn = get_db_connection()

        origen = conn.execute(
            """
            SELECT *
            FROM inventario
            WHERE producto_id = ?
            AND sucursal_id = ?
            """,
            (p_id, s_origen)
        ).fetchone()

        if not origen:
            return jsonify({
                "status": "error",
                "message": "No existe inventario del producto en la sucursal de origen."
            }), 404

        if origen['stock_actual'] < cantidad:
            return jsonify({
                "status": "error",
                "message": "Stock insuficiente en la sucursal de origen.",
                "stock_disponible": origen['stock_actual']
            }), 400

        destino = conn.execute(
            """
            SELECT *
            FROM inventario
            WHERE producto_id = ?
            AND sucursal_id = ?
            """,
            (p_id, s_destino)
        ).fetchone()

        sucursal_destino_base = conn.execute(
            """
            SELECT sucursal_uid
            FROM inventario
            WHERE sucursal_id = ?
            LIMIT 1
            """,
            (s_destino,)
        ).fetchone()

        if destino:
            dest_uid = destino['sucursal_uid']
        elif sucursal_destino_base:
            dest_uid = sucursal_destino_base['sucursal_uid']
        else:
            dest_uid = sanitizar_xss(str(data.get('sucursal_destino_uid', f"SUC-{s_destino}")).strip())

        if es_invalido(dest_uid):
            dest_uid = f"SUC-{s_destino}"

        orig_ant = origen['stock_actual']
        orig_nue = orig_ant - cantidad

        conn.execute(
            """
            UPDATE inventario
            SET stock_actual = ?,
                actualizado_en = datetime('now')
            WHERE id = ?
            """,
            (orig_nue, origen['id'])
        )

        if destino:
            dest_ant = destino['stock_actual']
            dest_nue = dest_ant + cantidad

            conn.execute(
                """
                UPDATE inventario
                SET stock_actual = ?,
                    actualizado_en = datetime('now')
                WHERE id = ?
                """,
                (dest_nue, destino['id'])
            )
        else:
            dest_ant = 0
            dest_nue = cantidad

            conn.execute(
                """
                INSERT INTO inventario
                (
                    producto_id,
                    producto_uid,
                    sucursal_id,
                    sucursal_uid,
                    stock_actual,
                    stock_reservado,
                    stock_minimo,
                    actualizado_en
                )
                VALUES (?, ?, ?, ?, ?, 0, 5, datetime('now'))
                """,
                (
                    p_id,
                    origen['producto_uid'],
                    s_destino,
                    dest_uid,
                    dest_nue
                )
            )

        # Movimiento de salida en sucursal origen
        conn.execute(
            """
            INSERT INTO movimientos_inventario
            (
                uid,
                producto_id,
                producto_uid,
                sucursal_id,
                sucursal_uid,
                tipo_movimiento,
                cantidad,
                stock_anterior,
                stock_nuevo,
                sucursal_origen_uid,
                sucursal_destino_uid,
                creado_en
            )
            VALUES (?, ?, ?, ?, ?, 'transferencia_salida', ?, ?, ?, ?, ?, datetime('now'))
            """,
            (
                str(uuid.uuid4()),
                p_id,
                origen['producto_uid'],
                s_origen,
                origen['sucursal_uid'],
                cantidad,
                orig_ant,
                orig_nue,
                origen['sucursal_uid'],
                dest_uid
            )
        )

        # Movimiento de entrada en sucursal destino
        conn.execute(
            """
            INSERT INTO movimientos_inventario
            (
                uid,
                producto_id,
                producto_uid,
                sucursal_id,
                sucursal_uid,
                tipo_movimiento,
                cantidad,
                stock_anterior,
                stock_nuevo,
                sucursal_origen_uid,
                sucursal_destino_uid,
                creado_en
            )
            VALUES (?, ?, ?, ?, ?, 'transferencia_entrada', ?, ?, ?, ?, ?, datetime('now'))
            """,
            (
                str(uuid.uuid4()),
                p_id,
                origen['producto_uid'],
                s_destino,
                dest_uid,
                cantidad,
                dest_ant,
                dest_nue,
                origen['sucursal_uid'],
                dest_uid
            )
        )

        conn.commit()

        disparar_evento_notificacion(
            "TransferCompleted",
            "Sistema Logística",
            f"Se transfirieron {cantidad} unidades desde {origen['sucursal_uid']} hacia {dest_uid}."
        )

        return jsonify({
            "status": "success",
            "message": "Transferencia completada exitosamente.",
            "origen": {
                "sucursal_id": s_origen,
                "sucursal_uid": origen['sucursal_uid'],
                "stock_anterior": orig_ant,
                "stock_actual": orig_nue
            },
            "destino": {
                "sucursal_id": s_destino,
                "sucursal_uid": dest_uid,
                "stock_anterior": dest_ant,
                "stock_actual": dest_nue
            }
        }), 200

    except Exception as e:
        if conn:
            conn.rollback()

        return jsonify({
            "status": "error",
            "message": "Error al procesar transferencia de inventario.",
            "detail": str(e)
        }), 500

    finally:
        if conn:
            conn.close()

    data = request.get_json() or {}
    for campo in ['producto_id', 'sucursal_origen_id', 'sucursal_destino_id', 'cantidad']:
        if campo not in data or es_invalido(data[campo]):
            return jsonify({"status": "error", "message": f"Falta el campo obligatorio: {campo}"}), 400

    cantidad = int(data['cantidad'])
    p_id = data['producto_id']
    s_origen = data['sucursal_origen_id']
    s_destino = data['sucursal_destino_id']

    try:
        conn = get_db_connection()
        origen = conn.execute("SELECT * FROM inventario WHERE producto_id = ? AND sucursal_id = ?", (p_id, s_origen)).fetchone()
        if not origen or origen['stock_actual'] < cantidad:
            conn.close()
            return jsonify({"status": "error", "message": "Stock insuficiente en la sucursal de origen."}), 400
            
        destino = conn.execute("SELECT * FROM inventario WHERE producto_id = ? AND sucursal_id = ?", (p_id, s_destino)).fetchone()
        
        orig_ant = origen['stock_actual']
        orig_nue = orig_ant - cantidad
        conn.execute("UPDATE inventario SET stock_actual = ?, actualizado_en = datetime('now') WHERE id = ?", (orig_nue, origen['id']))
        
        if destino:
            dest_ant = destino['stock_actual']
            dest_nue = dest_ant + cantidad
            conn.execute("UPDATE inventario SET stock_actual = ?, actualizado_en = datetime('now') WHERE id = ?", (dest_nue, destino['id']))
            dest_uid = destino['sucursal_uid']
        else:
            dest_ant = 0
            dest_nue = cantidad
            dest_uid = str(uuid.uuid4())
            conn.execute("INSERT INTO inventario (producto_id, producto_uid, sucursal_id, sucursal_uid, stock_actual, stock_reservado, stock_minimo, actualizado_en) VALUES (?, ?, ?, ?, ?, 0, 5, datetime('now'))", (p_id, origen['producto_uid'], s_destino, dest_uid, dest_nue))

        conn.execute("INSERT INTO movimientos_inventario (uid, producto_id, producto_uid, sucursal_id, sucursal_uid, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, sucursal_origen_uid, sucursal_destino_uid, creado_en) VALUES (?, ?, ?, ?, ?, 'transferencia_salida', ?, ?, ?, ?, ?, datetime('now'))", (str(uuid.uuid4()), p_id, origen['producto_uid'], s_origen, origen['sucursal_uid'], cantidad, orig_ant, orig_nue, origen['sucursal_uid'], dest_uid))
        conn.commit()
        conn.close()
        
        disparar_evento_notificacion("TransferCompleted", "Sistema Logística", f"Se transfirieron {cantidad} unidades desde sucursal ID {s_origen} a la {s_destino}.")
        return jsonify({"status": "success", "message": "Transferencia completada exitosamente."}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/inventory/loadExcel', methods=['POST'])
def load_excel():
    conn = None

    try:
        if 'file' not in request.files:
            return jsonify({
                "status": "error",
                "message": "Debe enviar un archivo Excel."
            }), 400

        archivo = request.files['file']

        if archivo.filename == '':
            return jsonify({
                "status": "error",
                "message": "No se seleccionó ningún archivo."
            }), 400

        nombre_archivo = archivo.filename.lower().strip()

        if not nombre_archivo.endswith('.xlsx'):
            return jsonify({
                "status": "error",
                "message": "Formato no válido. Solo se permite archivo Excel .xlsx."
            }), 400

        workbook = load_workbook(archivo, data_only=True)
        hoja = workbook.active

        if hoja.max_row < 2:
            return jsonify({
                "status": "error",
                "message": "El archivo Excel no contiene filas de inventario."
            }), 400

        conn = get_db_connection()

        filas_procesadas = 0
        filas_omitidas = 0
        errores = []

        # Estructura esperada:
        # Columna A: producto_id
        # Columna B: producto_uid
        # Columna C: sucursal_id
        # Columna D: sucursal_uid
        # Columna E: cantidad
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):

            producto_id_raw = fila[0] if len(fila) > 0 else None
            producto_uid_raw = fila[1] if len(fila) > 1 else None
            sucursal_id_raw = fila[2] if len(fila) > 2 else None
            sucursal_uid_raw = fila[3] if len(fila) > 3 else None
            cantidad_raw = fila[4] if len(fila) > 4 else None

            # Si la fila está completamente vacía, se omite.
            if (
                es_invalido(producto_id_raw)
                and es_invalido(producto_uid_raw)
                and es_invalido(sucursal_id_raw)
                and es_invalido(sucursal_uid_raw)
                and es_invalido(cantidad_raw)
            ):
                filas_omitidas += 1
                continue

            producto_id, error_producto = convertir_entero_positivo(producto_id_raw, 'producto_id')
            if error_producto:
                errores.append({
                    "fila": numero_fila,
                    "campo": "producto_id",
                    "message": error_producto
                })
                continue

            sucursal_id, error_sucursal = convertir_entero_positivo(sucursal_id_raw, 'sucursal_id')
            if error_sucursal:
                errores.append({
                    "fila": numero_fila,
                    "campo": "sucursal_id",
                    "message": error_sucursal
                })
                continue

            cantidad, error_cantidad = convertir_entero_positivo(cantidad_raw, 'cantidad')
            if error_cantidad:
                errores.append({
                    "fila": numero_fila,
                    "campo": "cantidad",
                    "message": error_cantidad
                })
                continue

            producto_uid = sanitizar_xss(str(producto_uid_raw).strip())
            sucursal_uid = sanitizar_xss(str(sucursal_uid_raw).strip())

            if es_invalido(producto_uid):
                errores.append({
                    "fila": numero_fila,
                    "campo": "producto_uid",
                    "message": "El campo producto_uid es obligatorio."
                })
                continue

            if es_invalido(sucursal_uid):
                errores.append({
                    "fila": numero_fila,
                    "campo": "sucursal_uid",
                    "message": "El campo sucursal_uid es obligatorio."
                })
                continue

            # Verificar que el producto exista en la tabla productos.
            producto = conn.execute(
                """
                SELECT id, uid
                FROM productos
                WHERE id = ?
                """,
                (producto_id,)
            ).fetchone()

            if not producto:
                errores.append({
                    "fila": numero_fila,
                    "campo": "producto_id",
                    "message": f"No existe un producto registrado con id {producto_id}."
                })
                continue

            registro = conn.execute(
                """
                SELECT *
                FROM inventario
                WHERE producto_id = ?
                AND sucursal_id = ?
                """,
                (producto_id, sucursal_id)
            ).fetchone()

            if registro:
                stock_anterior = registro['stock_actual']
                stock_nuevo = stock_anterior + cantidad

                conn.execute(
                    """
                    UPDATE inventario
                    SET stock_actual = ?,
                        actualizado_en = datetime('now')
                    WHERE id = ?
                    """,
                    (stock_nuevo, registro['id'])
                )

            else:
                stock_anterior = 0
                stock_nuevo = cantidad

                conn.execute(
                    """
                    INSERT INTO inventario
                    (
                        producto_id,
                        producto_uid,
                        sucursal_id,
                        sucursal_uid,
                        stock_actual,
                        stock_reservado,
                        stock_minimo,
                        actualizado_en
                    )
                    VALUES (?, ?, ?, ?, ?, 0, 5, datetime('now'))
                    """,
                    (
                        producto_id,
                        producto_uid,
                        sucursal_id,
                        sucursal_uid,
                        stock_nuevo
                    )
                )

            conn.execute(
                """
                INSERT INTO movimientos_inventario
                (
                    uid,
                    producto_id,
                    producto_uid,
                    sucursal_id,
                    sucursal_uid,
                    tipo_movimiento,
                    cantidad,
                    stock_anterior,
                    stock_nuevo,
                    referencia_tipo,
                    referencia_uid,
                    creado_en
                )
                VALUES
                (
                    ?, ?, ?, ?, ?,
                    'entrada_compra',
                    ?, ?, ?,
                    'importacion_excel',
                    ?,
                    datetime('now')
                )
                """,
                (
                    str(uuid.uuid4()),
                    producto_id,
                    producto_uid,
                    sucursal_id,
                    sucursal_uid,
                    cantidad,
                    stock_anterior,
                    stock_nuevo,
                    f"EXCEL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                )
            )

            filas_procesadas += 1

        if errores:
            conn.rollback()

            return jsonify({
                "status": "error",
                "message": "El archivo contiene errores. No se importó ninguna fila.",
                "filas_procesadas": 0,
                "filas_omitidas": filas_omitidas,
                "errores": errores
            }), 400

        if filas_procesadas == 0:
            conn.rollback()

            return jsonify({
                "status": "error",
                "message": "No se encontró ninguna fila válida para importar.",
                "filas_procesadas": 0,
                "filas_omitidas": filas_omitidas
            }), 400

        conn.commit()

        disparar_evento_notificacion(
            "InventoryLoaded",
            "Admin Logística",
            f"Se importaron {filas_procesadas} filas desde Excel."
        )

        return jsonify({
            "status": "success",
            "message": "Excel importado correctamente.",
            "filas_procesadas": filas_procesadas,
            "filas_omitidas": filas_omitidas
        }), 200

    except Exception as e:
        if conn:
            conn.rollback()

        return jsonify({
            "status": "error",
            "message": "Error al importar el archivo Excel.",
            "detail": str(e)
        }), 500

    finally:
        if conn:
            conn.close()

@app.route('/inventory/movements', methods=['GET'])
def get_movements():
    try:
        conn = get_db_connection()

        movimientos = conn.execute("""
            SELECT 
                m.id,
                m.uid,
                m.producto_id,
                m.producto_uid,
                p.codigo AS producto_codigo,
                p.nombre AS producto_nombre,
                m.sucursal_id,
                m.sucursal_uid,
                m.tipo_movimiento,
                m.cantidad,
                m.stock_anterior,
                m.stock_nuevo,
                m.referencia_tipo,
                m.referencia_id,
                m.referencia_uid,
                m.sucursal_origen_uid,
                m.sucursal_destino_uid,
                m.creado_en
            FROM movimientos_inventario m
            LEFT JOIN productos p ON p.id = m.producto_id
            ORDER BY datetime(m.creado_en) DESC, m.id DESC
        """).fetchall()

        conn.close()

        return jsonify({
            "status": "success",
            "total": len(movimientos),
            "data": [dict(m) for m in movimientos]
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": "Error al obtener los movimientos de inventario.",
            "detail": str(e)
        }), 500
    try:
        conn = get_db_connection()

        movimientos = conn.execute("""
            SELECT *
            FROM movimientos_inventario
            ORDER BY creado_en DESC
        """).fetchall()

        conn.close()

        return jsonify({
            "status": "success",
            "data": [dict(m) for m in movimientos]
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    
if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5002, debug=True)